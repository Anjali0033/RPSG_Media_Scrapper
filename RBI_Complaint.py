import requests
from lxml import html
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def scrape_disposal_table(url):
    response = requests.get(url)
    tree = html.fromstring(response.content)

    title_xpath = "//td[@class='head' and contains(text(),'Mode of disposal of Maintainable Complaints against Scheduled Commercial Banks')]"
    title_element = tree.xpath(title_xpath)
    title = title_element[0].text_content().strip() if title_element else "Title not found"

    base_xpath = "//td[@class='head' and contains(text(),'Mode of disposal of Maintainable Complaints against Scheduled Commercial Banks')]/ancestor::table[1]"
    rows = tree.xpath(f"{base_xpath}//tr")

    data = []
    for row in rows:
        cells = row.xpath('./td')
        row_data = [cell.text_content().strip() for cell in cells]
        if row_data:
            data.append(row_data)

    columns = data[1]
    data = data[2:]
    data = [row for row in data if len(row) == len(columns)]

    return title, columns, data

def save_to_excel(title, columns, data, filename='RBI_BANK_COMPLAINT.xlsx'):
    aligned_data = []
    for row in data:
        if len(row) > len(columns):
            aligned_data.append(row[:len(columns)])
        else:
            aligned_data.append(row + [''] * (len(columns) - len(row)))

    df = pd.DataFrame(aligned_data, columns=columns)

    second_col = columns[1]  # Total complaints column
    df[second_col] = df[second_col].astype(str).str.replace(',', '', regex=False)
    df[second_col] = pd.to_numeric(df[second_col], errors='coerce')
    df = df.dropna(subset=[second_col])

    # --- Add Total Income (₹ Cr) for each bank(Taking static values ref - https://www.screener.in/)
    income_data = {
        "STATE BANK OF INDIA": 386500,
        "ICICI BANK LIMITED": 142891,
        "HDFC BANK LIMITED": 198180,
        "AXIS BANK LIMITED": 121500,
        "PUNJAB NATIONAL BANK": 31895,
        "BANK OF BARODA": 32570,
        "UNION BANK OF INDIA": 27135,
        "CANARA BANK": 30751,
        "KOTAK MAHINDRA BANK LIMITED": 16633,
        "BANK OF INDIA": 15770,
        "INDIAN BANK": 15770,
        "CENTRAL BANK OF INDIA": 7112,
        "INDUSIND BANK LIMITED": 8381,
        "IDFC FIRST BANK LIMITED": 8381,
        "RBL BANK LIMITED": 8381,
        "UCO BANK": 7112,
        "YES BANK LIMITED": 8381,
        "STANDARD CHARTERED BANK": 8381,
        "BANK OF MAHARASHTRA": 7112,
        "INDIAN OVERSEAS BANK": 7112,
        "IDBI BANK LIMITED": 7819,
        "FEDERAL BANK LIMITED": 8381,
        "CITIBANK N.A": 8381,
        "SBM BANK (INDIA) LIMITED": 8381,
        "AU SMALL FINANCE BANK LIMITED": 8381,
        "BANDHAN BANK LIMITED": 8381,
        "PUNJAB AND SIND BANK": 7112,
        "KARUR VYSYA BANK LIMITED": 8381,
        "JAMMU & KASHMIR BANK LIMITED": 8381,
        "KARNATAKA BANK LIMITED": 8381,
        "SOUTH INDIAN BANK LIMITED": 8381,
        "DCB BANK LIMITED": 8381,
        "DBS BANK INDIA LIMITED": 8381,
        "AMERICAN EXPRESS BANKING CORP.": 8381,
        "HONGKONG AND SHANGHAI BANKING CORPN. LIMITED": 8381,
        "TAMILNAD MERCANTILE BANK LIMITED": 8381,
        "CITY UNION BANK LIMITED": 8381,
        "CSB BANK LIMITED": 8381,
        "DEUTSCHE BANK AG": 8381,
        "NAINITAL BANK LIMITED": 8381,
        "DHANLAXMI BANK LIMITED": 8381,
        "BARCLAYS BANK PLC": 8381,
        "BANK OF AMERICA NATIONAL ASSOCIATION": 8381,
        "WOORI BANK": 8381,
        "BNP PARIBAS": 8381,
        "ABU DHABI COMMERCIAL BANK PJSC": 8381,
        "JPMORGAN CHASE BANK NATIONAL ASSOCIATION": 8381,
        "SBER BANK": 8381,
        "NATWEST MARKETS PLC (ERSTWHILE THE ROYAL BANK OF SCOTLAND PLC)": 8381,
        "MUFG BANK LIMITED": 8381,
        "SHINHAN BANK": 8381,
    }

    df["Total Income (₹ Cr)"] = df[columns[0]].map(income_data)
    df["Complaints per ₹100 Cr Total Income"] = (df[second_col] / df["Total Income (₹ Cr)"] * 100).round(2)

    df_sorted = df.sort_values(by=second_col, ascending=False)

    # Update columns for Excel export
    updated_columns = list(df_sorted.columns)
    title_row = pd.DataFrame([[title] + [''] * (len(updated_columns) - 1)], columns=updated_columns)
    header_row = pd.DataFrame([updated_columns], columns=updated_columns)
    final_df = pd.concat([title_row, header_row, df_sorted], ignore_index=True)

    # Save to Excel
    final_df.to_excel(filename, index=False, header=False)

    wb = load_workbook(filename)
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(updated_columns))
    cell = ws.cell(row=1, column=1)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(updated_columns) + 1):
        max_length = max(len(str(ws.cell(row=row, column=col_idx).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    wb.save(filename)

# Run the scraper
url = 'https://www.rbi.org.in/Scripts/PublicationsView.aspx?id=22432'
title, columns, data = scrape_disposal_table(url)
save_to_excel(title, columns, data)
